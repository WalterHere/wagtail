import datetime
from io import BytesIO

from django.conf import settings
from django.contrib.admin.utils import quote
from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from django.urls import reverse
from django.utils.formats import date_format
from django.utils.timezone import make_aware
from openpyxl import load_workbook

from wagtail.models import ModelLogEntry
from wagtail.test.testapp.models import FeatureCompleteToy, JSONStreamModel
from wagtail.test.utils.template_tests import AdminTemplateTestUtils
from wagtail.test.utils.wagtail_tests import WagtailTestUtils
from wagtail.utils.deprecation import RemovedInWagtail60Warning


class TestModelViewSetGroup(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    def test_menu_items(self):
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        # Menu label falls back to the title-cased app label
        self.assertContains(
            response,
            '"name": "tests", "label": "Tests", "icon_name": "folder-open-inverse"',
        )
        # Title-cased from verbose_name_plural
        self.assertContains(response, "Json Stream Models")
        self.assertContains(response, reverse("streammodel:index"))
        self.assertEqual(reverse("streammodel:index"), "/admin/streammodel/")
        # Set on class
        self.assertContains(response, "JSON MinMaxCount StreamModel")
        self.assertContains(response, reverse("minmaxcount_streammodel:index"))
        self.assertEqual(
            reverse("minmaxcount_streammodel:index"),
            "/admin/minmaxcount-streammodel/",
        )
        # Set on instance
        self.assertContains(response, "JSON BlockCounts StreamModel")
        self.assertContains(response, reverse("blockcounts_streammodel:index"))
        self.assertEqual(
            reverse("blockcounts_streammodel:index"),
            "/admin/blockcounts/streammodel/",
        )


class TestTemplateConfiguration(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.default = JSONStreamModel.objects.create(
            body='[{"type": "text", "value": "foo"}]',
        )
        cls.custom = FeatureCompleteToy.objects.create(name="Test Toy")

    def get_default_url(self, view_name, args=()):
        return reverse(f"streammodel:{view_name}", args=args)

    def get_custom_url(self, view_name, args=()):
        return reverse(f"feature_complete_toy:{view_name}", args=args)

    def test_default_templates(self):
        pk = quote(self.default.pk)
        cases = {
            "index": (
                [],
                "wagtailadmin/generic/index.html",
            ),
            "index_results": (
                [],
                "wagtailadmin/generic/listing_results.html",
            ),
            "add": (
                [],
                "wagtailadmin/generic/create.html",
            ),
            "edit": (
                [pk],
                "wagtailadmin/generic/edit.html",
            ),
            "delete": (
                [pk],
                "wagtailadmin/generic/confirm_delete.html",
            ),
        }
        for view_name, (args, template_name) in cases.items():
            with self.subTest(view_name=view_name):
                response = self.client.get(self.get_default_url(view_name, args=args))
                self.assertTemplateUsed(response, template_name)

    def test_custom_template_lookups(self):
        pk = quote(self.custom.pk)
        cases = {
            "override with index_template_name": (
                "index",
                [],
                "tests/fctoy_index.html",
            ),
            "with app label and model name": (
                "add",
                [],
                "customprefix/tests/featurecompletetoy/create.html",
            ),
            "with app label": (
                "edit",
                [pk],
                "customprefix/tests/edit.html",
            ),
            "without app label and model name": (
                "delete",
                [pk],
                "customprefix/confirm_delete.html",
            ),
        }
        for case, (view_name, args, template_name) in cases.items():
            with self.subTest(case=case):
                response = self.client.get(self.get_custom_url(view_name, args=args))
                self.assertTemplateUsed(response, template_name)
                self.assertContains(
                    response, "<p>Some extra custom content</p>", html=True
                )

    def test_wagtail_admin_template_mixin_variables(self):
        pk = quote(self.custom.pk)
        cases = {
            "index": ([], "Feature complete toys", None),
            "add": ([], "New", "Feature complete toy"),
            "edit": ([pk], "Editing", str(self.custom)),
            "delete": ([pk], "Delete", str(self.custom)),
        }
        for view_name, (args, title, subtitle) in cases.items():
            with self.subTest(view_name=view_name):
                response = self.client.get(self.get_custom_url(view_name, args=args))
                soup = self.get_soup(response.content)
                h1 = soup.select_one("h1")
                self.assertIsNotNone(h1)
                self.assertEqual(
                    "".join(h1.find_all(string=True, recursive=False)).strip(), title
                )
                subtitle_el = h1.select_one("span")
                if subtitle:
                    self.assertIsNotNone(subtitle_el)
                    self.assertEqual(subtitle_el.string, subtitle)
                else:
                    self.assertIsNone(subtitle_el)
                icon = h1.select_one("svg use[href='#icon-media']")
                self.assertIsNotNone(icon)


class TestCustomColumns(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(name="Racecar")
        FeatureCompleteToy.objects.create(name="level")
        FeatureCompleteToy.objects.create(name="Lotso")

    def test_list_display(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url)
        # "name" column
        self.assertContains(response, "Racecar")
        self.assertContains(response, "level")
        self.assertContains(response, "Lotso")
        # BooleanColumn("is_cool")
        soup = self.get_soup(response.content)

        help = soup.select_one("td:has(svg.icon-help)")
        self.assertIsNotNone(help)
        self.assertEqual(help.text.strip(), "None")

        success = soup.select_one("td:has(svg.icon-success.w-text-positive-100)")
        self.assertIsNotNone(success)
        self.assertEqual(success.text.strip(), "True")

        error = soup.select_one("td:has(svg.icon-error.w-text-critical-100)")
        self.assertIsNotNone(error)
        self.assertEqual(error.text.strip(), "False")

        updated_at = soup.select("th a")[-1]
        self.assertEqual(updated_at.text.strip(), "Updated")
        self.assertEqual(updated_at["href"], f"{index_url}?ordering=_updated_at")


class TestListFilter(WagtailTestUtils, TestCase):
    cases = {
        "list": ("feature_complete_toy", "release_date", "Release date"),
        "dict": ("fctoy_alt1", "name__icontains", "Name contains"),
        "filterset_class": (
            "fctoy-alt2",
            "release_date__year__lte",
            "Release date year is less than or equal to",
        ),
    }

    def setUp(self):
        self.user = self.login()

    def get(self, url_namespace, params=None):
        return self.client.get(reverse(f"{url_namespace}:index"), params)

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Buzz Lightyear",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="Forky",
            release_date=datetime.date(2019, 6, 11),
        )

    def test_unfiltered_no_results(self):
        FeatureCompleteToy.objects.all().delete()
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace)
                self.assertTemplateUsed(response, "wagtailadmin/shared/filters.html")
                self.assertContains(
                    response,
                    "There are no feature complete toys to display",
                )
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.text.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)

    def test_unfiltered_with_results(self):
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace)
                self.assertTemplateUsed(response, "wagtailadmin/shared/filters.html")
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")
                self.assertNotContains(
                    response,
                    "There are no feature complete toys to display",
                )
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.text.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)

    def test_empty_filter_with_results(self):
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace, {lookup: ""})
                self.assertTemplateUsed(response, "wagtailadmin/shared/filters.html")
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.text.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertFalse(input.attrs.get("value"))

    def test_filtered_no_results(self):
        lookup_values = {
            "release_date": "1999-09-09",
            "name__icontains": "Woody",
            "release_date__year__lte": "1990",
        }
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {lookup: value})
                self.assertTemplateUsed(response, "wagtailadmin/shared/filters.html")
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.text.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertEqual(input.attrs.get("value"), value)

    def test_filtered_with_results(self):
        lookup_values = {
            "release_date": "1995-11-19",
            "name__icontains": "Ightyear",
            "release_date__year__lte": "2017",
        }
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {lookup: value})
                self.assertTemplateUsed(response, "wagtailadmin/shared/filters.html")
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertNotContains(response, "Forky")
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.text.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertEqual(input.attrs.get("value"), value)


class TestSearchIndexView(WagtailTestUtils, TestCase):
    url_name = "index"
    cases = {
        # With the default search backend
        "default": ("feature_complete_toy", "release_date"),
        # With Django ORM
        None: ("fctoy-alt2", "release_date__year__lte"),
    }

    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Buzz Lightyear",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="Forky",
            release_date=datetime.date(2019, 6, 11),
        )

    def assertInputRendered(self, response, search_q):
        soup = self.get_soup(response.content)
        input = soup.select_one("input#id_q")
        self.assertIsNotNone(input)
        self.assertEqual(input.attrs.get("value"), search_q)

    def get(self, url_namespace, params=None):
        return self.client.get(reverse(f"{url_namespace}:{self.url_name}"), params)

    def test_search_disabled(self):
        response = self.get("fctoy_alt1", {"q": "ork"})
        self.assertContains(response, "Forky")
        self.assertContains(response, "Buzz Lightyear")
        self.assertNotContains(response, "There are 2 matches")
        soup = self.get_soup(response.content)
        input = soup.select_one("input#id_q")
        self.assertIsNone(input)

    def test_search_no_results(self):
        for backend, (url_namespace, _) in self.cases.items():
            with self.subTest(backend=backend):
                response = self.get(url_namespace, {"q": "Woody"})
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertInputRendered(response, "Woody")

    def test_search_with_results(self):
        for backend, (url_namespace, _) in self.cases.items():
            with self.subTest(backend=backend):
                response = self.get(url_namespace, {"q": "ork"})
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertInputRendered(response, "ork")

    def test_filtered_searched_no_results(self):
        lookup_values = {
            "release_date": "2019-06-11",
            "release_date__year__lte": "2023",
        }
        for backend, (url_namespace, lookup) in self.cases.items():
            with self.subTest(backend=backend):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {"q": "Woody", lookup: value})
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertInputRendered(response, "Woody")

    def test_filtered_searched_with_results(self):
        lookup_values = {
            "release_date": "2019-06-11",
            "release_date__year__lte": "2023",
        }
        for backend, (url_namespace, lookup) in self.cases.items():
            with self.subTest(backend=backend):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {"q": "ork", lookup: value})
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertInputRendered(response, "ork")


class TestSearchIndexResultsView(TestSearchIndexView):
    url_name = "index_results"

    def assertInputRendered(self, response, search_q):
        # index_results view doesn't render the search input
        pass


class TestListExport(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Racecar",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="LEVEL",
            release_date=datetime.date(2010, 6, 18),
        )
        FeatureCompleteToy.objects.create(
            name="Catso",
            release_date=datetime.date(2010, 6, 18),
        )

    def test_export_disabled(self):
        index_url = reverse("fctoy_alt1:index")
        response = self.client.get(index_url)
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(f"a[href='{index_url}?export=csv']")
        self.assertIsNone(csv_link)
        xlsx_link = soup.select_one(f"a[href='{index_url}?export=xlsx']")
        self.assertIsNone(xlsx_link)

    def test_get_not_export_shows_export_buttons(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url)
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(f"a[href='{index_url}?export=csv']")
        self.assertIsNotNone(csv_link)
        self.assertEqual(csv_link.text.strip(), "Download CSV")
        xlsx_link = soup.select_one(f"a[href='{index_url}?export=xlsx']")
        self.assertIsNotNone(xlsx_link)
        self.assertEqual(xlsx_link.text.strip(), "Download XLSX")

    def test_get_filtered_shows_export_buttons_with_filters(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"release_date": "2010-06-18"})
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(
            f"a[href='{index_url}?release_date=2010-06-18&export=csv']"
        )
        self.assertIsNotNone(csv_link)
        self.assertEqual(csv_link.text.strip(), "Download CSV")
        xlsx_link = soup.select_one(
            f"a[href='{index_url}?release_date=2010-06-18&export=xlsx']"
        )
        self.assertIsNotNone(xlsx_link)
        self.assertEqual(xlsx_link.text.strip(), "Download XLSX")

    def test_csv_export(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"export": "csv"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.csv"',
        )

        data_lines = response.getvalue().decode().strip().split("\r\n")
        self.assertEqual(data_lines[0], "Name,Launch date,Is cool")
        self.assertEqual(data_lines[1], "Catso,2010-06-18,False")
        self.assertEqual(data_lines[2], "LEVEL,2010-06-18,True")
        self.assertEqual(data_lines[3], "Racecar,1995-11-19,")
        self.assertEqual(len(data_lines), 4)

    def test_csv_export_filtered(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(
            index_url,
            {"release_date": "2010-06-18", "export": "csv"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.csv"',
        )

        data_lines = response.getvalue().decode().strip().split("\r\n")
        self.assertEqual(data_lines[0], "Name,Launch date,Is cool")
        self.assertEqual(data_lines[1], "Catso,2010-06-18,False")
        self.assertEqual(data_lines[2], "LEVEL,2010-06-18,True")
        self.assertEqual(len(data_lines), 3)

    def test_xlsx_export(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"export": "xlsx"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.xlsx"',
        )

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data)).active
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]
        self.assertEqual(cell_array[0], ["Name", "Launch date", "Is cool"])
        self.assertEqual(cell_array[1], ["Catso", datetime.date(2010, 6, 18), False])
        self.assertEqual(cell_array[2], ["LEVEL", datetime.date(2010, 6, 18), True])
        self.assertEqual(cell_array[3], ["Racecar", datetime.date(1995, 11, 19), None])
        self.assertEqual(len(cell_array), 4)

    def test_xlsx_export_filtered(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(
            index_url,
            {"release_date": "2010-06-18", "export": "xlsx"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.xlsx"',
        )

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data)).active
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]
        self.assertEqual(cell_array[0], ["Name", "Launch date", "Is cool"])
        self.assertEqual(cell_array[1], ["Catso", datetime.date(2010, 6, 18), False])
        self.assertEqual(cell_array[2], ["LEVEL", datetime.date(2010, 6, 18), True])
        self.assertEqual(len(cell_array), 3)


class TestPagination(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        objects = [FeatureCompleteToy(name=f"Frisbee {i}") for i in range(32)]
        FeatureCompleteToy.objects.bulk_create(objects)

    def test_default_list_pagination(self):
        list_url = reverse("fctoy_alt1:index")
        response = self.client.get(list_url)

        # Default is 20 per page
        self.assertEqual(FeatureCompleteToy.objects.all().count(), 32)
        self.assertContains(response, "Page 1 of 2")
        self.assertContains(response, "Next")
        self.assertContains(response, list_url + "?p=2")

    def test_custom_list_pagination(self):
        list_url = reverse("feature_complete_toy:index")
        response = self.client.get(list_url)

        # Custom is set to display 5 per page
        self.assertEqual(FeatureCompleteToy.objects.all().count(), 32)
        self.assertContains(response, "Page 1 of 7")
        self.assertContains(response, "Next")
        self.assertContains(response, list_url + "?p=2")


class TestOrdering(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        objects = [
            FeatureCompleteToy(name="CCCCCCCCCC", strid="1"),
            FeatureCompleteToy(name="AAAAAAAAAA", strid="2"),
            FeatureCompleteToy(name="DDDDDDDDDD", strid="3"),
            FeatureCompleteToy(name="BBBBBBBBBB", strid="4"),
        ]
        FeatureCompleteToy.objects.bulk_create(objects)

    def test_default_order(self):
        response = self.client.get(reverse("fctoy_alt1:index"))
        # Without ordering on the model, should be ordered descending by pk
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "BBBBBBBBBB",
                "DDDDDDDDDD",
                "AAAAAAAAAA",
                "CCCCCCCCCC",
            ],
        )

    def test_custom_order(self):
        response = self.client.get(reverse("feature_complete_toy:index"))
        # Should respect the viewset's ordering
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "AAAAAAAAAA",
                "BBBBBBBBBB",
                "CCCCCCCCCC",
                "DDDDDDDDDD",
            ],
        )


class TestBreadcrumbs(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")

    def test_index_view(self):
        response = self.client.get(reverse("feature_complete_toy:index"))
        items = [
            {
                "url": "",
                "label": "Feature complete toys",
            }
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_add_view(self):
        response = self.client.get(reverse("feature_complete_toy:add"))
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": "",
                "label": "New: Feature complete toy",
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_edit_view(self):
        edit_url = reverse("feature_complete_toy:edit", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": "",
                "label": str(self.object),
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_delete_view(self):
        delete_url = reverse(
            "feature_complete_toy:delete",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(delete_url)
        self.assertBreadcrumbsNotRendered(response.content)

    def test_history_view(self):
        history_url = reverse(
            "feature_complete_toy:history",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(history_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": reverse(
                    "feature_complete_toy:edit", args=(quote(self.object.pk),)
                ),
                "label": str(self.object),
            },
            {
                "url": "",
                "label": "History",
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)


class TestLegacyPatterns(WagtailTestUtils, TestCase):
    # RemovedInWagtail60Warning: legacy integer pk-based URLs will be removed

    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = JSONStreamModel.objects.create(
            body='[{"type": "text", "value": "foo"}]',
        )

    def test_legacy_edit(self):
        edit_url = reverse("streammodel:edit", args=(quote(self.object.pk),))
        legacy_edit_url = "/admin/streammodel/1/"
        with self.assertWarnsRegex(
            RemovedInWagtail60Warning,
            "`/<pk>/` edit view URL pattern has been deprecated in favour of /edit/<pk>/.",
        ):
            response = self.client.get(legacy_edit_url)
        self.assertEqual(edit_url, "/admin/streammodel/edit/1/")
        self.assertRedirects(response, edit_url, 301)

    def test_legacy_delete(self):
        delete_url = reverse("streammodel:delete", args=(quote(self.object.pk),))
        legacy_delete_url = "/admin/streammodel/1/delete/"
        with self.assertWarnsRegex(
            RemovedInWagtail60Warning,
            "`/<pk>/delete/` delete view URL pattern has been deprecated in favour of /delete/<pk>/.",
        ):
            response = self.client.get(legacy_delete_url)
        self.assertEqual(delete_url, "/admin/streammodel/delete/1/")
        self.assertRedirects(response, delete_url, 301)


class TestHistoryView(WagtailTestUtils, TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = cls.create_test_user()
        cls.object = FeatureCompleteToy.objects.create(name="Buzz")
        cls.url = reverse(
            "feature_complete_toy:history",
            args=(quote(cls.object.pk),),
        )

        content_type = ContentType.objects.get_for_model(FeatureCompleteToy)
        cls.timestamp_1 = datetime.datetime(2021, 9, 30, 10, 1, 0)
        cls.timestamp_2 = datetime.datetime(2022, 5, 10, 12, 34, 0)
        if settings.USE_TZ:
            cls.timestamp_1 = make_aware(cls.timestamp_1)
            cls.timestamp_2 = make_aware(cls.timestamp_2)
        ModelLogEntry.objects.create(
            content_type=content_type,
            label="Test Buzz",
            action="wagtail.create",
            user=cls.user,
            timestamp=cls.timestamp_1,
            object_id=cls.object.pk,
        )
        ModelLogEntry.objects.create(
            content_type=content_type,
            label="Test Buzz Updated",
            action="wagtail.edit",
            user=cls.user,
            timestamp=cls.timestamp_2,
            object_id=cls.object.pk,
        )

    def setUp(self):
        self.login(self.user)

    def test_simple(self):
        expected = (
            ("Edited", str(self.user), date_format(self.timestamp_2, "c")),
            ("Created", str(self.user), date_format(self.timestamp_1, "c")),
        )
        response = self.client.get(self.url)
        soup = self.get_soup(response.content)
        rows = soup.select("tbody tr")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 2)

        rendered_rows = []
        for row in rows:
            cells = []
            tds = row.select("td")
            self.assertEqual(len(tds), 3)
            cells.append(tds[0].text.strip())
            cells.append(tds[1].text.strip())
            cells.append(tds[2].select_one("time").attrs.get("datetime"))
            rendered_rows.append(cells)

        for rendered_row, expected_row in zip(rendered_rows, expected):
            self.assertSequenceEqual(rendered_row, expected_row)

    def test_filters(self):
        response = self.client.get(self.url, {"action": "wagtail.edit"})
        soup = self.get_soup(response.content)
        rows = soup.select("tbody tr")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].select_one("td").text.strip(), "Edited")

        response = self.client.get(self.url, {"action": "wagtail.create"})
        soup = self.get_soup(response.content)
        rows = soup.select("tbody tr")
        heading = soup.select_one("h2")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(heading.text.strip(), "There is 1 match")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].select_one("td").text.strip(), "Created")

    def test_empty(self):
        ModelLogEntry.objects.all().delete()
        response = self.client.get(self.url)
        soup = self.get_soup(response.content)
        results = soup.select_one("#listing-results")
        table = soup.select_one("table")
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(results)
        self.assertEqual(results.text.strip(), "No log entries found.")
        self.assertIsNone(table)

    def test_edit_view_links_to_history_view(self):
        edit_url = reverse("feature_complete_toy:edit", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        soup = self.get_soup(response.content)
        header = soup.select_one(".w-slim-header")
        history_link = header.find("a", attrs={"href": self.url})
        self.assertIsNotNone(history_link)
